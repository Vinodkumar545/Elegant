
from openpyxl import Workbook

work_book = Workbook()

sheet_name = 'testdata'

work_book.create_sheet(sheet_name)

work_sheet = work_book[sheet_name]

work_sheet.cell(row=1, column=1).value = 'TC_ID'
work_sheet.cell(row=1, column=2).value = 'TC_Name'
work_sheet.cell(row=1, column=3).value = 'TC_Run Mode'


work_sheet.cell(row=2, column=1).value = 'TC_001'
work_sheet.cell(row=2, column=2).value = 'Test cases xxxxxx'
work_sheet.cell(row=2, column=3).value = 'Yes'


work_book.save("new_wb.xlsx")
