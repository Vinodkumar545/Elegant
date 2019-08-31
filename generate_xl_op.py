
from openpyxl import load_workbook

work_book = load_workbook('elegant_data1.xlsx')

sheet_name = 'testdata'

col_tcid, col_tcname, col_runmode = 1, 2, 3
col_username, col_password, col_url = 4, 5, 6

row_start = 2

# to get number of rows
tot_rows = len(list(work_book[sheet_name].columns)[0])
print(tot_rows)

result = {}
for row in range(row_start, tot_rows + 1):

	# read cell value for the column run mode
	run_mode = work_book[sheet_name].cell(row=row, column=col_runmode).value

	if run_mode.lower() == 'yes':

		tc_para = {}

		tc_id = work_book[sheet_name].cell(row=row, column=col_tcid).value
		tc_name = work_book[sheet_name].cell(row=row, column=col_tcname).value
		username = work_book[sheet_name].cell(row=row, column=col_username).value
		password = work_book[sheet_name].cell(row=row, column=col_password).value
		url = work_book[sheet_name].cell(row=row, column=col_url).value

		tc_para['username'] = username
		tc_para['password'] = password
		tc_para['url'] = url

		result[tc_id] = tc_para

		print(tc_id, tc_name, run_mode, username, password, url)


print(result)

import json

with open('test_data_xl.json', 'w') as writefile:
	json.dump(result, writefile)





