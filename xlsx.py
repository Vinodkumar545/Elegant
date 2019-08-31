
from openpyxl import load_workbook
from openpyxl import Workbook
import os


excel_name = 'elegant_data1.xlsx'
print(os.path.join(os.getcwd(), excel_name))

excel_path = os.path.join(os.getcwd(), excel_name)

print(os.path.exists(excel_path))

# load existing work book
work_book = load_workbook(excel_path)

# to get all the sheetnames in the workbook
print(work_book.sheetnames)

# to create sheet in the workbook
# work_book.create_sheet("testdata_sheet", 0)
# print(work_book.sheetnames)

sheet_name = 'testdata'

# # read a cell
# print(work_book[sheet_name].cell(row=2, column=2).value)
if work_book[sheet_name].cell(row=10, column=2).value is None:
	print("cell value is None")

# # # write into cell
# # work_book[sheet_name].cell(row=4, column=1).value = 'TC003'
# # work_book[sheet_name].cell(row=4, column=2).value = 'Test Case execution scenario 3'
# # work_book[sheet_name].cell(row=4, column=3).value = "Yes"

# always save the excel
work_book.save(excel_name)

# to get number of rows
tot_rows = len(list(work_book[sheet_name].columns)[0])
print(tot_rows)

# to get number of columns
tot_cols = len(list(work_book[sheet_name].rows)[0])
print(tot_cols)


col_tcid, col_tcname, col_runmode = 1, 2, 3

for row in range(2, tot_rows + 1):

	# read cell value for the column run mode
	run_mode = work_book[sheet_name].cell(row=row, column=col_runmode).value

	tc_id = work_book[sheet_name].cell(row=row, column=col_tcid).value

	tc_name = work_book[sheet_name].cell(row=row, column=col_tcname).value

	print(tc_id, tc_name, run_mode)


import xlsx_base

print(xlsx_base.row_list(work_book, sheet_name, "Find row of this cell (4, 2)", 2))


# create a new work book


