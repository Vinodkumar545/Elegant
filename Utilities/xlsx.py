from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
from openpyxl.comments import Comment
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl import Workbook
import logging
from Utilities import settings


LOGGER = logging.getLogger(settings.LOG_NAME)

def load_excel(file_path):
	"""
	To activate the workbook and as a argument send excel file path
	"""
	try:
		work_book = load_workbook(filename=file_path, data_only=True) # data_only helps read just cell value not formula in the cell
		LOGGER.info("Workbook opened!")
		return work_book
	except Exception as e:
		LOGGER.exception("active_excel Exception: %s", e)

def create_workbook():
	"""
	To create a new workbook (By default 'Sheet1' would be created)
	"""
	try:
		wb = Workbook()
		LOGGER.info("New Workbook Created!")
		return wb
	except Exception as e:
		LOGGER.info("create_workbook Exception: %s", e)

def create_sheet(workbook, list_of_sheet_names):
	"""
	To create new sheets
	Args:
		workbook: workbook object that was created 
		list_of_sheet_names: provide list of sheet names to be created
	"""
	try:
		for sheet_index, sheet_name in enumerate(list_of_sheet_names):
			workbook.create_sheet(sheet_name, sheet_index)

		LOGGER.info("List of Sheet provided are created")
	except Exception as e:
		LOGGER.exception("create_sheet Exception: %s", e)

def get_all_sheet_name(work_book):
	"""
	To get all the sheet name from a work book. Returns list of sheetnames
	"""
	try:
		sheet_names = work_book.sheetnames
		log_str = "No. of sheet names in the work book is : " + str(len(sheet_names)) + " - " + str(sheet_names)
		LOGGER.info(log_str)
		return sheet_names
	except Exception as e:
		LOGGER.exception("get_all_sheet_name Exception: %s",e)

def read_cell(work_book, sheet_name, row_num, col_num, log_status=True):
	"""
	To read a cell value from a particular sheet of a workbook. Returns the cell values
	In below condition, if the cell is empty, then None is returned. Change the condition as per the need
	"""
	try:
		work_sheet = work_book[sheet_name]
		cell_value = work_sheet.cell(row=row_num, column=col_num).value
		if cell_value is not None and log_status == True: # If cell is empty/null, None is returned. so, we are not logging such values.
			log_str = "The value '" + str(cell_value) + "' is successfully read"
			LOGGER.info(log_str)
		return cell_value
	except Exception as e:
		LOGGER.exception("read_cell Exception: %s", e)

def write_cell(work_book, sheet_name, row_num, col_num, value, log_status=True):
	"""
	To wrtie a value in a call on a particular sheet of a workbook
	"""
	try:
		work_sheet = work_book[sheet_name]
		work_sheet.cell(row=row_num, column=col_num).value = value
		if log_status:
			log_str = "The value '" + str(value) + "' is successfully written!"
			LOGGER.info(log_str)
	except Exception as e:
		LOGGER.exception("write_cell Exception: %s", e)

def get_row_length(work_book, sheet_name, log_status=True):
	"""
	This functiona return total number of rows which has cell values. 
	Note: Its not fixed for no. of rows for particalur column but instead it considers all the columns
	E.g., If Col A as 25 rows and Col B has 35 rows filled with cell values, then '35' is returned
	"""
	try:
		work_sheet = work_book[sheet_name]
		row_len = len(list(work_sheet.columns)[0])
		if log_status:
			LOGGER.info("Total number of rows are: %s", row_len)
		return row_len
	except Exception as e:
		LOGGER.exception("get_row_length Exception: %s", e)

def get_column_length(work_book, sheet_name, log_status=True):
	"""
	To find number of columns filled with values from a particualar sheet of a workbook. Returns total column length (integer)
	"""
	try:
		work_sheet = work_book[sheet_name]
		col_len = len(list(work_sheet.rows)[1])
		if log_status:
			LOGGER.info("Total number of columns are: %s", col_len)
		return col_len
	except Exception as e:
		LOGGER.exception("get_column_length Exception: %s", e)

def save_excel(work_book, save_name):
	"""
	To save a excel workbook, please don't forget to call this functionality!
	"""
	try:
		work_book.save(save_name)
		LOGGER.info("Save Workbook!")
	except Exception as e:
		LOGGER.exception("save_excel Exception: %s", e)

def row_list(work_book, sheet_name, value, col_num):
	"""
	To find row number of particular cell value. Returns row number which matching values in the form of list, in case of multipe values
	"""
	try:
		work_sheet = work_book[sheet_name]
		row_no = []

		# LOGGER.info("%s", list(work_sheet.columns)[col_num-1])
		for obj in list(work_sheet.columns)[col_num-1]:
			if obj.value is None:
				break

			if obj.value == value:
				row_no.append(obj.row)
				# break

			# if str(value) in obj.value:
			# 	row_no.append(obj.row)
			# 	# break

			# elif obj.value in str(value):
			# 	row_no.append(obj.row)
			# 	# break

		if len(row_no) == 0:
			log_str = "The value '" + str(value) + "' is not present in any row in " + str(sheet_name) 
			LOGGER.info(log_str)

		else:
			log_str = "The value '" + str(value) + "' is present in the row no. " + str(row_no) + " in " + str(sheet_name)
			LOGGER.info(log_str)
		return row_no
	except Exception as e:
		LOGGER.exception("row_list Exception: %s",e)